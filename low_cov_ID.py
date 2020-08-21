from openpyxl import load_workbook
import xlsxwriter

###
def ret_content_file(HLA):
    wb=load_workbook("HLA-"+HLA+".xlsx")
    ws=wb['Sheet1']
    all_rows=list(ws.rows)
    content=[]
    for i in range(len(all_rows)):
        row=[]
        for cell in all_rows[i]:
            row.append(cell.value)
        content.append(row)
    CONTENT=[]
    for i in range(len(content[1])):
        CONTENT.append([content[1][i]])
    for i in range(2,len(content)-8): #to remove the pos and neg controls
        for j in range(len(content[i])):
            CONTENT[j].append(content[i][j])
    if HLA in ["DQA1","DQA1F"]:
        CONTENT=CONTENT[:-3]
    if HLA in ["DQB1","DQB1F","DRB1","DRB1F"]:
        CONTENT=CONTENT[:-2]
    return CONTENT

def ret_content_file_2_3(HLA):
    wb=load_workbook("HLA-"+HLA+".xlsx")
    ws=wb['Sheet1']
    all_rows=list(ws.rows)
    content=[]
    for i in range(len(all_rows)):
        row=[]
        for cell in all_rows[i]:
            row.append(cell.value)
        content.append(row)
    CONTENT=[]
    for i in range(len(content[1])):
        CONTENT.append([content[1][i]])
    content=content[1:]
    for i in range(2,len(content)-8):
        for j in range(len(content[i])):
            CONTENT[j].append(content[i][j])
    if HLA in ["DQA1","DQA1F","DQB1","DQB1F","DRB1","DRB1F"]:
        CONTENT=[CONTENT[0]]+[CONTENT[2]]
    if HLA in ["A","AF","B","BF","C","CF"]:
        CONTENT=[CONTENT[0]]+[CONTENT[2]]+[CONTENT[3]]
    return CONTENT
  
def get_min_ID(list):
    Lid,Lval=[],[]
    for i in range(1,len(list)):
        subLid,subLval=[],[]
        L=sorted(list[i])
        val=L[17]       #first quartile
        for j in range(len(list[i])):
            if list[i][j]<=val:
                subLid.append(list[0][j])
                subLval.append(list[i][j])
        Lid.append(subLid)
        Lval.append(subLval)
    return Lid, Lval

def how_many(Lid, HLA):
    count={}
    for i in range(len(Lid)):
        for j in range(len(Lid[i])):
            if Lid[i][j] in count: 
                count[Lid[i][j]]=count[Lid[i][j]]+1
            else:
                count[Lid[i][j]]=1
    return count


    
def get_ID():
    ID=[]
    id=[]
    with open('list_ID.txt','r') as list_ID:
        for i in list_ID.readlines():
            ID=ID+[i.rstrip('\n')]
    ID=ID[:-8]
    for i in ID:
        id.append([i])
    return id
    
def complete(glob,id,hla):
    
    for HLA in hla:
        lilg=glob[HLA]
        K=[]
        for key in lilg:
            K.append(key)
        for j in range(len(id)):
            if id[j][0] in K:
                id[j].append(lilg[id[j][0]])
            else:
                id[j].append(0)
    return id
        

def output_csv_mean(id, hla,name):
    name="low_cov_list"+name+".xlsx"
    workbook = xlsxwriter.Workbook(name)
    worksheet = workbook.add_worksheet()
    bold = workbook.add_format({'bold': True})
    # Header
    row = 0
    col = 1
    for HLA in hla:
        worksheet.write(row, col,HLA,bold)
        col+=1
    #1st column
    row=1
    col=0
    for sample in id:
        worksheet.write(row,col,sample[0],bold)
        row+=1
    #Data
    row=1
    col=1
    for sample in id: 
        for val in sample[1:]: 
            worksheet.write(row,col,val) #####
            col+=1
        row+=1
        col=1
    workbook.close()
    

####
# Count the nb of time a sample is in the lowest 25% for each exon
def main():
    hla=['A','AF','B', 'BF', 'C', 'CF', 'DQA1', 'DQA1F', 'DQB1', 'DQB1F', 'DRB1', 'DRB1F']
    glob={}
    for HLA in hla:
        l=ret_content_file(HLA)
        Lid, Lval=get_min_ID(l)
        count=how_many(Lid, HLA)
        glob[HLA]=count
    id=get_ID()
    id=complete(glob,id,hla)
    output_csv_mean(id,hla,'')
    
# Count the nb of time a sample is in the lowest 25% on exon 2 and 3 for class I and exon 2 for class II
def main_exon23():
    hla=['A','AF','B', 'BF', 'C', 'CF', 'DQA1', 'DQA1F', 'DQB1', 'DQB1F', 'DRB1', 'DRB1F']
    glob={}
    for HLA in hla:
        l=ret_content_file_2_3(HLA)
        Lid, Lval=get_min_ID(l)
        count=how_many(Lid, HLA)
        glob[HLA]=count
    id=get_ID()
    id=complete(glob,id,hla)
    output_csv_mean(id,hla,'ex2_3')
    


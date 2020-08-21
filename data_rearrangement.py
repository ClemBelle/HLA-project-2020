from openpyxl import load_workbook
import xlsxwriter


#### Get the list of samples ID 
def get_ID():
    ID=[]
    with open('list_ID.txt','r') as list_ID:
        for i in list_ID.readlines():
            ID=ID+[i.rstrip('\n')]
    return ID


#### Retrieve data from the excels
def ret_content_file(name):
    wb=load_workbook(name)
    ws=wb['Sheet1']
    all_rows=list(ws.rows)
    content=[]
    for i in range(len(all_rows)):
        row=[]
        for cell in all_rows[i]:
            row.append(cell.value)
        content.append(row)
    content=content[1:]
    content=[content[0][1:]]+[content[2][1:]]+[content[4][1:]]+[content[6][1:]]+[content[8][1:]]+[content[10][1:]]+[content[12][1:]]+[content[14][1:]]
    content[5]=content[5][:6]+[-1, -1]+content[5][8:]
    content[6]=content[6][:6]+[-1, -1,-1, -1,-1, -1]
    content[7]=content[7][:6]+[-1, -1,-1, -1,-1, -1]
    a=[]
    af=[]
    b=[]
    bf=[]
    c=[]
    cf=[]
    dqa=[]
    dqaf=[]
    dqb=[]
    dqbf=[]
    drb=[]
    drbf=[]
    
    for i in content:
        
        for j in range(len(i)):
            if j==0:
                a.append(i[j])
            elif j==1:
                af.append(i[j])
            elif j==2:
                b.append(i[j])
            elif j==3:
                bf.append(i[j])
            elif j==4:
                c.append(i[j])
            elif j==5:
                cf.append(i[j])
            elif j==6:
                dqa.append(i[j])
            elif j==7:
                dqaf.append(i[j])
            elif j==8:
                dqb.append(i[j])
            elif j==9:
                dqbf.append(i[j])
            elif j==10:
                drb.append(i[j])
            elif j==11:
                drbf.append(i[j])
    return a,af,b,bf,c,cf,dqa,dqaf,dqb,dqbf,drb,drbf

def get_means(ID,name):
    name=ID[0]+name
    a,af,b,bf,c,cf,dqa,dqaf,dqb,dqbf,drb,drbf=ret_content_file(name)
    a=[[a[0]]]+[[a[1]]]+[[a[2]]]+[[a[3]]]+[[a[4]]]+[[a[5]]]+[[a[6]]]+[[a[7]]]
    af=[[af[0]]]+[[af[1]]]+[[af[2]]]+[[af[3]]]+[[af[4]]]+[[af[5]]]+[[af[6]]]+[[af[7]]]
    b=[[b[0]]]+[[b[1]]]+[[b[2]]]+[[b[3]]]+[[b[4]]]+[[b[5]]]+[[b[6]]]+[[b[7]]]
    bf=[[bf[0]]]+[[bf[1]]]+[[bf[2]]]+[[bf[3]]]+[[bf[4]]]+[[bf[5]]]+[[bf[6]]]+[[bf[7]]]
    c=[[c[0]]]+[[c[1]]]+[[c[2]]]+[[c[3]]]+[[c[4]]]+[[c[5]]]+[[c[6]]]+[[c[7]]]
    cf=[[cf[0]]]+[[cf[1]]]+[[cf[2]]]+[[cf[3]]]+[[cf[4]]]+[[cf[5]]]+[[cf[6]]]+[[cf[7]]]
    dqa=[[dqa[0]]]+[[dqa[1]]]+[[dqa[2]]]+[[dqa[3]]]+[[dqa[4]]]+[[dqa[5]]]+[[dqa[6]]]+[[dqa[7]]]
    dqaf=[[dqaf[0]]]+[[dqaf[1]]]+[[dqaf[2]]]+[[dqaf[3]]]+[[dqaf[4]]]+[[dqaf[5]]]+[[dqaf[6]]]+[[dqaf[7]]]
    dqb=[[dqb[0]]]+[[dqb[1]]]+[[dqb[2]]]+[[dqb[3]]]+[[dqb[4]]]+[[dqb[5]]]+[[dqb[6]]]+[[dqb[7]]]
    dqbf=[[dqbf[0]]]+[[dqbf[1]]]+[[dqbf[2]]]+[[dqbf[3]]]+[[dqbf[4]]]+[[dqbf[5]]]+[[dqbf[6]]]+[[dqbf[7]]]
    drb=[[drb[0]]]+[[drb[1]]]+[[drb[2]]]+[[drb[3]]]+[[drb[4]]]+[[drb[5]]]+[[drb[6]]]+[[drb[7]]]
    drbf=[[drbf[0]]]+[[drbf[1]]]+[[drbf[2]]]+[[drbf[3]]]+[[drbf[4]]]+[[drbf[5]]]+[[drbf[6]]]+[[drbf[7]]]
    
    for i in range(1,len(ID)):
        name='_coverage_means_HLA_genes.xlsx'
        name=ID[i]+name
        A,AF,B,BF,C,CF,DQA,DQAF,DQB,DQBF,DRB,DRBF=ret_content_file(name)
        for i in range(len(A)):
            a[i].append(A[i])
            af[i].append(AF[i])
            b[i].append(B[i])
            bf[i].append(BF[i])
            c[i].append(C[i])
            cf[i].append(CF[i])
            dqa[i].append(DQA[i])
            dqaf[i].append(DQAF[i])
            dqb[i].append(DQB[i])
            dqbf[i].append(DQBF[i])
            drb[i].append(DRB[i])
            drbf[i].append(DRBF[i])
                
    return a,af,b,bf,c,cf,dqa,dqaf,dqb,dqbf,drb,drbf
    


#### Create the new excel
def output_csv_mean(ID, L, name):
    """Creates a .csv file to store the mean coverage of each segments of the HLA genes"""
    header=[]
    for i in range(len(L)):
        header.append("exon"+str(i+1))
    
    name=name+".xlsx"
    workbook = xlsxwriter.Workbook(name)
    worksheet = workbook.add_worksheet()
    bold = workbook.add_format({'bold': True})
    # Header
    row = 0
    col = 1
    for j in range(len(L)):
        worksheet.write(row, col,header[j],bold)
        col+=1
    #1st column
    row=1
    col=0
    for name in ID:
        worksheet.write(row,col,name,bold)
        row+=1
    #Data
    row=1
    col=1
    for exon in L: 
        for mean in exon: 
            worksheet.write(row,col,mean) #####
            row+=1
        col+=1
        row=1
    workbook.close()
    
### MAIN 
ID=get_ID()
name='_coverage_means_HLA_genes.xlsx'
a,af,b,bf,c,cf,dqa,dqaf,dqb,dqbf,drb,drbf=get_means(ID,name)
output_csv_mean(ID, a, "HLA-A")
output_csv_mean(ID, af, "HLA-AF")
output_csv_mean(ID, b, "HLA-B")
output_csv_mean(ID, bf, "HLA-BF")
output_csv_mean(ID, c, "HLA-C")
output_csv_mean(ID, cf, "HLA-CF")
output_csv_mean(ID, dqa, "HLA-DQA1")
output_csv_mean(ID, dqaf, "HLA-DQA1F")
output_csv_mean(ID, dqb, "HLA-DQB1")
output_csv_mean(ID, dqbf, "HLA-DQB1F")
output_csv_mean(ID, drb, "HLA-DRB1")
output_csv_mean(ID, drbf, "HLA-DRB1F")
